import requests
import openpyxl
import time
import random
import csv
import os
from bs4 import BeautifulSoup
from openpyxl.styles import Font, Alignment, PatternFill

# --- CONFIG ---
BASE_URL = "https://books.toscrape.com/catalogue/page-{}.html"
OUTPUT_FILE = "Scraped_Books_Data.xlsx"
BATCH_SIZE = 10  
RATING_MAP = {'One': 1, 'Two': 2, 'Three': 3, 'Four': 4, 'Five': 5}
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
}

def fetch_data(url):
    for attempt in range(3):
        try:
            response = requests.get(url, headers=HEADERS, timeout=10)
            response.encoding = "utf-8"
            if response.status_code == 200:
                return response.text
            print(f"[-] Attempt {attempt + 1} failed. Status: {response.status_code}")
        except requests.exceptions.RequestException as e:
            print(f"[-] Attempt {attempt + 1} error: {e}")
        time.sleep(2) 
    return None

def parse_data(html):
    soup = BeautifulSoup(html, 'html.parser')
    books_on_page = []
    
    for book in soup.find_all('article', class_='product_pod'):
        title = book.find('h3').find('a')['title']
        rating_class = book.find(class_='star-rating')['class'][1]
        rating = RATING_MAP.get(rating_class, 0)
        
        price_text = book.find(class_='price_color').text.strip()
        try:
            price = float(price_text.replace('£', ''))
        except ValueError:
            price = 0.0

        availability = book.find(class_='availability').text.strip()
        books_on_page.append((title, price, availability, rating))
    
    has_next = soup.find('li', class_='next') is not None
    return books_on_page, has_next

def save_to_excel(data):
    if os.path.exists(OUTPUT_FILE):
        workbook = openpyxl.load_workbook(OUTPUT_FILE)
        sheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Books Catalog"
        
        headers = ["Title", "Price (£)", "Availability", "Rating (1-5)"]
        header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.fill = header_fill
        
        sheet.column_dimensions["A"].width = 55
        sheet.column_dimensions["B"].width = 12
        sheet.column_dimensions["C"].width = 18
        sheet.column_dimensions["D"].width = 12

    for book in data:
        sheet.append(book)

    try:
        workbook.save(OUTPUT_FILE)
    except PermissionError:
        print(f"[!] Close '{OUTPUT_FILE}' to save the current session!")

def save_to_csv(data):
    file_exists = os.path.exists("Scraped_Books_Data.csv")
    with open("Scraped_Books_Data.csv", mode='a', encoding='utf-8', newline='') as file:
        writer = csv.writer(file)
        if not file_exists:
            writer.writerow(["Title", "Price (£)", "Availability", "Rating (1-5)"])
        writer.writerows(data)

def main():
    page = 1
    total_books = 0
    keep_running = True
    
    print(f"Starting Scraper (Session size: {BATCH_SIZE} pages)")
    
    if os.path.exists(OUTPUT_FILE): os.remove(OUTPUT_FILE)
    if os.path.exists("Scraped_Books_Data.csv"): os.remove("Scraped_Books_Data.csv")

    while keep_running:
        session_books = []
        
        for _ in range(BATCH_SIZE):
            print(f"Scraping Page {page}...")
            url = BASE_URL.format(page)
            html = fetch_data(url)
            
            if not html:
                keep_running = False
                break
                
            books, has_next = parse_data(html)
            session_books.extend(books)
            
            page += 1 

            if not has_next:
                print("Reached the final page.")
                keep_running = False
                break 
                
            time.sleep(random.uniform(0.5, 1.5))

        if session_books:
            total_books += len(session_books)
            save_to_excel(session_books)
            save_to_csv(session_books)
            print(f"Checkpoint saved! Pages processed: {page-1} | Total Books: {total_books}")
        
        if keep_running:
            print("Session cooldown (5 seconds)...")
            time.sleep(5)

    print(f"\n Finished! {total_books} books saved to {OUTPUT_FILE} and CSV")
    
if __name__ == "__main__":
    main()